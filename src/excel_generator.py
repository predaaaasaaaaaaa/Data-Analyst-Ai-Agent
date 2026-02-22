import logging
import pandas as pd
from pathlib import Path
from typing import Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ExcelReportGenerator:
    def __init__(self):
        self.logger = logger
        
    def generate_report(self, df: pd.DataFrame, analysis: Dict[str, Any], output_path: str) -> Optional[str]:
        """
        Generate professional Excel report with analysis results
        
        Args:
            df: Original DataFrame
            analysis: Analysis results from DataAnalyzer
            output_path: Path to save Excel file
            
        Returns:
            Path to generated Excel file or None if failed
        """
        try:
            output_path = Path(output_path)
            output_path.parent.mkdir(parents=True, exist_ok=True)
            
            self.logger.info(f"Generating Excel report: {output_path}")
            
            # Create workbook
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            # Add sheets
            self._add_data_sheet(wb, df)
            self._add_summary_stats_sheet(wb, analysis)
            self._add_data_quality_sheet(wb, analysis)
            self._add_correlations_sheet(wb, analysis)
            self._add_outliers_sheet(wb, analysis)
            self._add_insights_sheet(wb, analysis)
            
            # Save
            wb.save(str(output_path))
            self.logger.info(f"Report generated successfully: {output_path}")
            
            return str(output_path)
            
        except Exception as e:
            self.logger.error(f"Error generating Excel report: {e}", exc_info=True)
            return None

    def _add_data_sheet(self, wb: Workbook, df: pd.DataFrame):
        """Add original data sheet"""
        ws = wb.create_sheet("Raw Data", 0)
        
        # Add headers
        for col_num, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = col_name
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add data
        for row_num, row in enumerate(df.values, 2):
            for col_num, value in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Adjust column widths
        for col_num in range(1, len(df.columns) + 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 20
        
        # Add filter
        ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}"

    def _add_summary_stats_sheet(self, wb: Workbook, analysis: Dict[str, Any]):
        """Add summary statistics sheet"""
        ws = wb.create_sheet("Summary Statistics", 1)
        
        stats = analysis.get('descriptive_stats', {})
        if not stats or 'note' in stats:
            ws['A1'] = "No numeric data available for statistics"
            return
        
        row = 1
        # Headers
        headers = ['Column', 'Mean', 'Median', 'Std Dev', 'Min', 'Max', 'Q25', 'Q75', 'Skewness']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        
        # Add statistics
        row = 2
        for col_name, col_stats in stats.items():
            if isinstance(col_stats, dict):
                ws.cell(row=row, column=1).value = col_name
                ws.cell(row=row, column=2).value = col_stats.get('mean', '')
                ws.cell(row=row, column=3).value = col_stats.get('median', '')
                ws.cell(row=row, column=4).value = col_stats.get('std', '')
                ws.cell(row=row, column=5).value = col_stats.get('min', '')
                ws.cell(row=row, column=6).value = col_stats.get('max', '')
                ws.cell(row=row, column=7).value = col_stats.get('q25', '')
                ws.cell(row=row, column=8).value = col_stats.get('q75', '')
                ws.cell(row=row, column=9).value = col_stats.get('skewness', '')
                row += 1
        
        # Format numbers
        for row in ws.iter_rows(min_row=2, max_row=row-1, min_col=2, max_col=9):
            for cell in row:
                cell.number_format = '0.00'
        
        # Adjust widths
        for col_num in range(1, 10):
            ws.column_dimensions[get_column_letter(col_num)].width = 15

    def _add_data_quality_sheet(self, wb: Workbook, analysis: Dict[str, Any]):
        """Add data quality sheet"""
        ws = wb.create_sheet("Data Quality", 2)
        
        quality = analysis.get('data_quality', {})
        
        row = 1
        ws.cell(row=row, column=1).value = "Data Quality Report"
        ws.cell(row=row, column=1).font = Font(bold=True, size=14)
        
        row += 2
        ws.cell(row=row, column=1).value = "Metric"
        ws.cell(row=row, column=2).value = "Value"
        for col in [1, 2]:
            ws.cell(row=row, column=col).font = Font(bold=True, color="FFFFFF")
            ws.cell(row=row, column=col).fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        
        row += 1
        ws.cell(row=row, column=1).value = "Duplicate Rows"
        ws.cell(row=row, column=2).value = quality.get('duplicate_rows', 0)
        row += 1
        
        ws.cell(row=row, column=1).value = "Duplicate %"
        ws.cell(row=row, column=2).value = f"{quality.get('duplicate_percentage', 0):.2f}%"
        row += 1
        
        missing = quality.get('missing_values', {})
        if missing:
            ws.cell(row=row, column=1).value = "Missing Values"
            ws.cell(row=row, column=2).value = sum(missing.values())
            row += 1
        
        data_types = quality.get('data_types', {})
        if data_types:
            ws.cell(row=row, column=1).value = "Numeric Columns"
            ws.cell(row=row, column=2).value = data_types.get('numeric', 0)
            row += 1
            
            ws.cell(row=row, column=1).value = "Categorical Columns"
            ws.cell(row=row, column=2).value = data_types.get('categorical', 0)
            row += 1
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20

    def _add_correlations_sheet(self, wb: Workbook, analysis: Dict[str, Any]):
        """Add correlations sheet"""
        ws = wb.create_sheet("Correlations", 3)
        
        corr = analysis.get('correlations', {})
        strong_corr = corr.get('strong_correlations', {})
        
        if not strong_corr:
            ws['A1'] = "No strong correlations found (threshold: 0.7)"
            return
        
        row = 1
        ws.cell(row=row, column=1).value = "Column Pair"
        ws.cell(row=row, column=2).value = "Correlation"
        for col in [1, 2]:
            ws.cell(row=row, column=col).font = Font(bold=True, color="FFFFFF")
            ws.cell(row=row, column=col).fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        
        row = 2
        for pair, corr_value in strong_corr.items():
            ws.cell(row=row, column=1).value = pair
            ws.cell(row=row, column=2).value = f"{corr_value:.3f}"
            ws.cell(row=row, column=2).number_format = '0.000'
            row += 1
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15

    def _add_outliers_sheet(self, wb: Workbook, analysis: Dict[str, Any]):
        """Add outliers sheet"""
        ws = wb.create_sheet("Outliers", 4)
        
        outliers = analysis.get('outliers', {})
        
        if 'note' in outliers:
            ws['A1'] = outliers['note']
            return
        
        if not outliers:
            ws['A1'] = "No outliers detected"
            return
        
        row = 1
        ws.cell(row=row, column=1).value = "Column"
        ws.cell(row=row, column=2).value = "Count"
        ws.cell(row=row, column=3).value = "Percentage"
        for col in [1, 2, 3]:
            ws.cell(row=row, column=col).font = Font(bold=True, color="FFFFFF")
            ws.cell(row=row, column=col).fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        
        row = 2
        for col_name, outlier_info in outliers.items():
            ws.cell(row=row, column=1).value = col_name
            ws.cell(row=row, column=2).value = outlier_info.get('count', 0)
            ws.cell(row=row, column=3).value = f"{outlier_info.get('percentage', 0):.2f}%"
            row += 1
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15

    def _add_insights_sheet(self, wb: Workbook, analysis: Dict[str, Any]):
        """Add insights and recommendations sheet"""
        ws = wb.create_sheet("Insights", 5)
        
        ws['A1'] = "Analysis Insights & Recommendations"
        ws['A1'].font = Font(bold=True, size=14)
        
        insights_data = analysis.get('insights', {})
        insights_list = insights_data.get('insights', [])
        
        row = 3
        for insight in insights_list:
            ws.cell(row=row, column=1).value = insight
            ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
            row += 1
        
        ws.column_dimensions['A'].width = 80
        for row_num in range(3, row):
            ws.row_dimensions[row_num].height = 30


def generate_report(df: pd.DataFrame, analysis: Dict[str, Any], output_path: str) -> Optional[str]:
    """Convenience function"""
    generator = ExcelReportGenerator()
    return generator.generate_report(df, analysis, output_path)
